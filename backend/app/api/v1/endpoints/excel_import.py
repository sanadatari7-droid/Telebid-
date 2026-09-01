from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.db.postgres import get_db, fetch_all, execute, fetch_val, require_company
from app.middleware.auth import require_roles, get_current_user
import io, json

router = APIRouter(prefix="/excel-import", tags=["Excel Import"])

@router.post("/analyze")
async def analyze_excel(file: UploadFile = File(...),
    conn=Depends(get_db), current_user=Depends(require_roles("ADMIN"))):
    """Analyze uploaded Excel file — return structure without importing."""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx/.xls) or CSV")
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            first_data_rows = []
            all_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"Col_{j}" for j, h in enumerate(row)]
                    continue
                if not any(cell is not None for cell in row):
                    continue
                str_row = [str(v) if v is not None else "" for v in row]
                if i <= 5:
                    first_data_rows.append(str_row)
                # Full dataset the frontend actually submits on confirm — capped as a
                # sanity limit, not a preview limit (sample_rows above is the preview).
                if len(all_rows) < 2000:
                    all_rows.append(dict(zip(headers, str_row)))
            sheets.append({
                "name": sheet_name,
                "headers": headers,
                "row_count": ws.max_row,
                "sample_rows": first_data_rows,
                "all_rows": all_rows,
                "suggested_mapping": _suggest_mapping(headers)
            })
        return {
            "file_name": file.filename,
            "sheets": sheets,
            "total_sheets": len(sheets),
            "analysis_complete": True,
            "message": "File analyzed. Review the structure and confirm import."
        }
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Add to requirements.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {str(e)}")

def _suggest_mapping(headers: list) -> dict:
    """Suggest what each column likely represents."""
    mapping = {}
    h_lower = [h.lower() for h in headers]
    hints = {
        "category": ["category", "type", "section", "group"],
        "field_name": ["field", "name", "question", "criterion", "item"],
        "dropdown_values": ["values", "options", "dropdown", "choices", "possible"],
        "weight": ["weight", "weighting", "%", "percentage"],
        "max_score": ["max", "maximum", "score", "points"],
        "min_score": ["min", "minimum", "threshold"],
        "mandatory": ["mandatory", "required", "compulsory", "must"],
        "description": ["description", "notes", "comment", "detail"],
    }
    for i, h in enumerate(h_lower):
        for field_type, keywords in hints.items():
            if any(kw in h for kw in keywords):
                mapping[headers[i]] = field_type
                break
        else:
            mapping[headers[i]] = "ignore"
    return mapping

@router.post("/import-evaluation-criteria")
async def import_eval_criteria(body: dict, conn=Depends(get_db),
    current_user=Depends(require_roles("ADMIN"))):
    """Import evaluation criteria from pre-analyzed Excel data (after user confirmation)."""
    company_id = require_company(current_user)
    rows = body.get("rows", [])
    mapping = body.get("column_mapping", {})
    template_name = body.get("template_name", "Imported Template")

    if not rows:
        raise HTTPException(status_code=400, detail="No rows provided")

    # Create import log
    import_id = await fetch_val(conn,
        """INSERT INTO excel_imports (company_id, file_name, import_type, total_rows, imported_by)
           VALUES ($1,$2,'EVALUATION_CRITERIA',$3,$4) RETURNING import_id""",
        company_id, template_name, len(rows), current_user.user_id)

    # Create evaluation template
    await execute(conn,
        "INSERT INTO evaluation_templates (tmpl_name, description, created_by, company_id) VALUES ($1,$2,$3,$4)",
        template_name, f"Imported from Excel on {__import__('datetime').datetime.now().strftime('%Y-%m-%d')}", current_user.user_id, company_id)
    tmpl_id = await fetch_val(conn,
        "SELECT tmpl_id FROM evaluation_templates WHERE company_id=$1 ORDER BY tmpl_id DESC LIMIT 1", company_id)

    imported = skipped = errors = 0
    error_details = []
    seen_categories = {}

    for i, row in enumerate(rows):
        try:
            cat_name = str(row.get(mapping.get("category",""), f"Category {i}") or f"Category {i}").strip()
            field_name = str(row.get(mapping.get("field_name",""), "") or "").strip()
            if not field_name:
                skipped += 1; continue

            weight_raw = row.get(mapping.get("weight",""), 0) or 0
            try:
                weight = float(str(weight_raw).replace('%','').strip())
            except: weight = 0

            max_score_raw = row.get(mapping.get("max_score",""), 100) or 100
            try:
                max_score = float(str(max_score_raw).strip())
            except: max_score = 100

            # Determine criteria type from category name
            cat_upper = cat_name.upper()
            if any(w in cat_upper for w in ["TECH","TECHNICAL"]):
                crit_type = "TECHNICAL"
            elif any(w in cat_upper for w in ["FIN","FINANCIAL","COMMERCIAL","PRICE"]):
                crit_type = "FINANCIAL"
            else:
                crit_type = "TECHNICAL"

            await execute(conn,
                """INSERT INTO evaluation_criteria (tmpl_id, crit_name, crit_type, weight, max_score, description, sort_order)
                   VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                tmpl_id, field_name, crit_type, weight, max_score,
                str(row.get(mapping.get("description",""), "") or ""),
                i)

            # Store dropdown values if present
            dropdown_col = mapping.get("dropdown_values","")
            dropdown_raw = str(row.get(dropdown_col, "") or "").strip()
            if dropdown_raw:
                options = [o.strip() for o in dropdown_raw.split('/') if o.strip()]
                for opt in options:
                    await execute(conn,
                        """INSERT INTO dropdown_configs (company_id, dropdown_key, dropdown_label, option_value, option_label, sort_order)
                           VALUES ($1,$2,$3,$4,$5,$6) ON CONFLICT DO NOTHING""",
                        company_id, f"eval_{tmpl_id}_{i}", field_name, opt.upper().replace(' ','_'), opt, options.index(opt))
            imported += 1
        except Exception as e:
            errors += 1
            error_details.append(f"Row {i+1}: {str(e)}")

    await execute(conn,
        "UPDATE excel_imports SET imported=$1, skipped=$2, errors=$3, status=$4, error_details=$5, completed_at=NOW() WHERE import_id=$6",
        imported, skipped, errors, "COMPLETED" if errors == 0 else "PARTIAL",
        json.dumps(error_details[:20]) if error_details else None, import_id)

    return {
        "import_id": import_id,
        "template_id": tmpl_id,
        "total": len(rows),
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "error_details": error_details[:10],
        "message": f"Import complete. {imported} criteria imported into template '{template_name}'"
    }

@router.get("/history")
async def import_history(conn=Depends(get_db), current_user=Depends(require_roles("ADMIN"))):
    company_id = require_company(current_user)
    return await fetch_all(conn, """
        SELECT ei.*, u.full_name AS imported_by_name
        FROM excel_imports ei LEFT JOIN users u ON ei.imported_by=u.user_id
        WHERE ei.company_id=$1
        ORDER BY ei.imported_at DESC LIMIT 50""", company_id)
